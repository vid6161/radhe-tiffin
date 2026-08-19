from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

from database import get_connection

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = Path(os.environ.get("RADHE_EXCEL_PATH", str(BASE_DIR / "radhe_tiffin_business.xlsx")))
EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="2D241E")
ACCENT_FILL = PatternFill("solid", fgColor="F4E2CF")
GOOD_FILL = PatternFill("solid", fgColor="E6F4EA")
WARN_FILL = PatternFill("solid", fgColor="FFF4CC")
BAD_FILL = PatternFill("solid", fgColor="FCE8E6")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=18, bold=True, color="2D241E")
SUBTITLE_FONT = Font(size=11, italic=True, color="6B5D53")
THIN = Side(style="thin", color="D9D2CC")
BORDER = Border(bottom=THIN)
CURRENCY = '€#,##0.00'
DATE_FMT = 'YYYY-MM-DD'


def _rows(connection, query, params=()):
    cursor = connection.cursor()
    cursor.execute(query, params)
    return cursor.fetchall()




def _value(row, key, default=None):
    """Safely read a sqlite3.Row field, allowing older databases/migrations.

    Some existing Radhe Tiffin databases may have been created before a
    reporting field was introduced. Excel generation should not fail merely
    because an optional reporting field is absent.
    """
    try:
        keys = row.keys()
        if key in keys:
            return row[key]
    except (AttributeError, TypeError, KeyError):
        pass
    return default

def _sheet_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUBTITLE_FONT


def _header(ws, row, headers):
    for col, value in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _finish_sheet(ws, freeze="A4", widths=None, autofilter=True):
    ws.freeze_panes = freeze
    if autofilter and ws.max_row >= 4 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    else:
        for idx in range(1, ws.max_column + 1):
            max_len = 0
            for cell in ws[get_column_letter(idx)]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 36)


def _currency_columns(ws, columns, start_row=4):
    for col in columns:
        for row in range(start_row, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = CURRENCY


def _date_columns(ws, columns, start_row=4):
    for col in columns:
        for row in range(start_row, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = DATE_FMT


def generate_excel_workbook(path: str | Path | None = None) -> Path:
    """Regenerate the Excel-compatible master business workbook from SQLite.

    This is intentionally a snapshot generated from the website's source of truth.
    It is safe to regenerate repeatedly and does not mutate business records.
    """
    output = Path(path) if path else EXCEL_PATH
    output.parent.mkdir(parents=True, exist_ok=True)

    connection = get_connection()
    try:
        orders = _rows(connection, """
            SELECT o.order_number, o.created_at, o.pickup_date, o.pickup_time,
                   o.pickup_location, o.quantity, o.total_amount, o.status,
                   o.payment_status, o.payment_method, o.payment_reference,
                   c.first_name, c.last_name, c.email, c.phone,
                   m.meal_name, m.day_name
            FROM orders o
            JOIN customers c ON c.id = o.customer_id
            JOIN menu m ON m.id = o.menu_id
            ORDER BY o.created_at DESC
        """)
        catering = _rows(connection, """
            SELECT r.request_number, r.created_at, r.event_date, r.package_type,
                   r.people_count, r.location, r.estimated_amount,
                   r.payment_status, r.status, c.first_name, c.last_name,
                   c.email, c.phone
            FROM catering_requests r
            JOIN customers c ON c.id = r.customer_id
            ORDER BY r.created_at DESC
        """)
        packages = _rows(connection, """
            SELECT p.package_number, p.created_at, p.frequency, p.weekly_price,
                   p.total_amount, p.pickup_location, p.pickup_time,
                   p.payment_method, p.payment_status, p.status,
                   c.first_name, c.last_name, c.email, c.phone
            FROM weekly_packages p
            JOIN customers c ON c.id = p.customer_id
            ORDER BY p.created_at DESC
        """)
        expenses = _rows(connection, """
            SELECT id, expense_date, category, description, vendor, amount, notes, created_at
            FROM expenses ORDER BY expense_date DESC, id DESC
        """)
        salaries = _rows(connection, """
            SELECT id, pay_date, employee_name, role, hours, hourly_rate,
                   gross_amount, notes, created_at
            FROM salaries ORDER BY pay_date DESC, id DESC
        """)
        inventory = _rows(connection, """
            SELECT id, transaction_date, item_name, transaction_type, quantity,
                   unit, unit_cost, total_cost, notes, created_at
            FROM inventory_transactions
            ORDER BY transaction_date DESC, id DESC
        """)
        menu = _rows(connection, """
            SELECT menu_date, day_name, day_name_german, meal_name, price, is_active
            FROM menu ORDER BY menu_date DESC, id DESC
        """)
        customers = _rows(connection, """
            SELECT id, first_name, last_name, email, phone, created_at
            FROM customers ORDER BY created_at DESC
        """)
        pickup = _rows(connection, """
            SELECT pickup_date, location, start_time, end_time, maximum_orders,
                   current_orders, is_active
            FROM pickup_slots ORDER BY pickup_date DESC, location, start_time
        """)
        settings = _rows(connection, "SELECT * FROM business_settings WHERE id=1")
        allergens = _rows(connection, """
            SELECT m.menu_date, m.day_name, m.meal_name, mi.item_name, mia.allergen_code, mia.notes
            FROM menu_item_allergens mia
            JOIN menu_items mi ON mi.id = mia.menu_item_id
            JOIN menu m ON m.id = mi.menu_id
            ORDER BY m.menu_date, mi.id, mia.allergen_code
        """)
    finally:
        connection.close()

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # ---------------- Dashboard ----------------
    ws = wb.create_sheet("Dashboard")
    _sheet_title(ws, "Radhe Tiffin — Business Dashboard", "Automatically generated from the website database")
    metrics = [
        ("Gross sales", sum(float(_value(r, "total_amount") or 0) for r in orders if _value(r, "status") != "cancelled"), CURRENCY),
        ("Collected sales", sum(float(_value(r, "total_amount") or 0) for r in orders if _value(r, "status") != "cancelled" and _value(r, "payment_status") == "paid"), CURRENCY),
        ("Outstanding sales", sum(float(_value(r, "total_amount") or 0) for r in orders if _value(r, "status") != "cancelled" and _value(r, "payment_status") != "paid"), CURRENCY),
        ("Other catering estimates", sum(float(_value(r, "estimated_amount") or 0) for r in catering if _value(r, "status") != "cancelled"), CURRENCY),
        ("Ingredient / usage cost", sum(float(_value(r, "total_cost") or 0) for r in inventory if str(_value(r, "transaction_type")).lower() in {"usage", "waste"}), CURRENCY),
        ("Other expenses", sum(float(_value(r, "amount") or 0) for r in expenses), CURRENCY),
        ("Salaries", sum(float(_value(r, "gross_amount") or 0) for r in salaries), CURRENCY),
    ]
    for i, (label, value, fmt) in enumerate(metrics, start=4):
        ws.cell(i, 1, label).font = Font(bold=True)
        ws.cell(i, 2, value).number_format = fmt
        ws.cell(i, 1).fill = ACCENT_FILL
        ws.cell(i, 2).fill = GOOD_FILL if "sales" in label.lower() else WARN_FILL
    ws["A12"] = "Operational profit / loss"
    ws["A12"].font = Font(bold=True, size=13)
    ws["B12"] = "=B4-B8-B9-B10"
    ws["B12"].number_format = CURRENCY
    ws["A12"].fill = ACCENT_FILL
    ws["B12"].fill = GOOD_FILL
    ws["A14"] = "Generated"
    ws["B14"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _finish_sheet(ws, freeze="A4", widths={"A": 30, "B": 22}, autofilter=False)

    # ---------------- Orders ----------------
    ws = wb.create_sheet("Orders")
    _sheet_title(ws, "Orders")
    headers = ["Order No.", "Created", "Pickup Date", "Pickup Time", "Location", "Qty", "Amount", "Status", "Payment Status", "Payment Method", "Payment Ref", "Customer", "Email", "Phone", "Meal", "Day"]
    _header(ws, 3, headers)
    for row_idx, r in enumerate(orders, 4):
        values = [_value(r, "order_number"), _value(r, "created_at"), _value(r, "pickup_date"), _value(r, "pickup_time"), _value(r, "pickup_location"), _value(r, "quantity"), _value(r, "total_amount"), _value(r, "status"), _value(r, "payment_status"), _value(r, "payment_method"), _value(r, "payment_reference"), f'{_value(r, "first_name")} {_value(r, "last_name")}', _value(r, "email"), _value(r, "phone"), _value(r, "meal_name"), _value(r, "day_name")]
        for c, v in enumerate(values, 1): ws.cell(row_idx, c, v)
    _currency_columns(ws, [7])
    _finish_sheet(ws, widths={"A": 22, "B": 20, "C": 14, "D": 20, "E": 14, "F": 8, "G": 14, "H": 14, "I": 16, "J": 18, "K": 28, "L": 24, "M": 28, "N": 18, "O": 26, "P": 14})

    # ---------------- Revenue ----------------
    ws = wb.create_sheet("Revenue")
    _sheet_title(ws, "Revenue", "Sales by pickup date; cancelled orders excluded")
    _header(ws, 3, ["Date", "Tiffins", "Gross Sales", "Paid", "Outstanding", "Cancelled Tiffins"])
    daily = defaultdict(lambda: {"qty": 0, "sales": 0.0, "paid": 0.0, "outstanding": 0.0, "cancelled": 0})
    for r in orders:
        d = _value(r, "pickup_date")
        qty = int(_value(r, "quantity") or 0)
        amount = float(_value(r, "total_amount") or 0)
        if _value(r, "status") == "cancelled":
            daily[d]["cancelled"] += qty
        else:
            daily[d]["qty"] += qty
            daily[d]["sales"] += amount
            if _value(r, "payment_status") == "paid": daily[d]["paid"] += amount
            else: daily[d]["outstanding"] += amount
    for row_idx, d in enumerate(sorted(daily), 4):
        x = daily[d]
        vals = [d, _value(x, "qty"), _value(x, "sales"), _value(x, "paid"), _value(x, "outstanding"), _value(x, "cancelled")]
        for c, v in enumerate(vals, 1): ws.cell(row_idx, c, v)
    _currency_columns(ws, [3,4,5])
    _finish_sheet(ws, widths={"A": 14, "B": 12, "C": 16, "D": 16, "E": 18, "F": 18})
    if ws.max_row > 4:
        chart = LineChart()
        chart.title = "Daily Sales"
        chart.y_axis.title = "EUR"
        chart.x_axis.title = "Date"
        data = Reference(ws, min_col=3, min_row=3, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, "H3")

    # ---------------- Expenses ----------------
    ws = wb.create_sheet("Expenses")
    _sheet_title(ws, "Expenses")
    _header(ws, 3, ["ID", "Date", "Category", "Description", "Vendor", "Amount", "Notes", "Created"])
    for i, r in enumerate(expenses, 4):
        vals = [_value(r, "id"), _value(r, "expense_date"), _value(r, "category"), _value(r, "description"), _value(r, "vendor"), _value(r, "amount"), _value(r, "notes"), _value(r, "created_at")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _currency_columns(ws,[6])
    _finish_sheet(ws, widths={"A":8,"B":14,"C":18,"D":34,"E":24,"F":14,"G":34,"H":20})

    # ---------------- Salaries ----------------
    ws = wb.create_sheet("Salaries")
    _sheet_title(ws, "Salaries")
    _header(ws, 3, ["ID", "Pay Date", "Employee", "Role", "Hours", "Hourly Rate", "Gross Amount", "Notes", "Created"])
    for i, r in enumerate(salaries, 4):
        vals = [_value(r, "id"), _value(r, "pay_date"), _value(r, "employee_name"), _value(r, "role"), _value(r, "hours"), _value(r, "hourly_rate"), _value(r, "gross_amount"), _value(r, "notes"), _value(r, "created_at")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _currency_columns(ws,[6,7])
    _finish_sheet(ws, widths={"A":8,"B":14,"C":24,"D":22,"E":12,"F":16,"G":18,"H":34,"I":20})

    # ---------------- Usage / Inventory ----------------
    ws = wb.create_sheet("Usage")
    _sheet_title(ws, "Inventory & Usage", "Purchase, usage, waste and stock adjustments")
    _header(ws, 3, ["ID", "Date", "Item", "Type", "Quantity", "Unit", "Unit Cost", "Total Cost", "Notes", "Created"])
    for i, r in enumerate(inventory, 4):
        vals = [_value(r, "id"), _value(r, "transaction_date"), _value(r, "item_name"), _value(r, "transaction_type"), _value(r, "quantity"), _value(r, "unit"), _value(r, "unit_cost"), _value(r, "total_cost"), _value(r, "notes"), _value(r, "created_at")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _currency_columns(ws,[7,8])
    _finish_sheet(ws, widths={"A":8,"B":14,"C":28,"D":16,"E":14,"F":12,"G":16,"H":16,"I":34,"J":20})

    # ---------------- P&L ----------------
    ws = wb.create_sheet("Profit & Loss")
    _sheet_title(ws, "Profit & Loss", "Usage/waste are treated as operational food cost; purchases remain inventory until used. Not a substitute for formal tax/accounting records")
    _header(ws, 3, ["Month", "Sales", "Ingredient / Usage Cost", "Other Expenses", "Salaries", "Operational Profit / Loss"])
    months = set()
    for r in orders:
        if _value(r, "pickup_date") and _value(r, "status") != "cancelled": months.add(str(_value(r, "pickup_date"))[:7])
    for r in expenses: months.add(str(_value(r, "expense_date"))[:7])
    for r in salaries: months.add(str(_value(r, "pay_date"))[:7])
    for r in inventory: months.add(str(_value(r, "transaction_date"))[:7])
    for row_idx, month in enumerate(sorted(months), 4):
        sales = sum(float(_value(r, "total_amount") or 0) for r in orders if _value(r, "status") != "cancelled" and str(_value(r, "pickup_date")).startswith(month))
        usage = sum(float(_value(r, "total_cost") or 0) for r in inventory if str(_value(r, "transaction_date")).startswith(month) and str(_value(r, "transaction_type")).lower() in {"usage", "waste"})
        other = sum(float(_value(r, "amount") or 0) for r in expenses if str(_value(r, "expense_date")).startswith(month))
        salary = sum(float(_value(r, "gross_amount") or 0) for r in salaries if str(_value(r, "pay_date")).startswith(month))
        vals = [month, sales, usage, other, salary, sales-usage-other-salary]
        for c,v in enumerate(vals,1): ws.cell(row_idx,c,v)
    _currency_columns(ws,[2,3,4,5,6])
    _finish_sheet(ws, widths={"A":14,"B":16,"C":24,"D":18,"E":16,"F":24})
    if ws.max_row > 4:
        chart = BarChart()
        chart.title = "Monthly Profit / Loss"
        chart.y_axis.title = "EUR"
        data = Reference(ws, min_col=6, min_row=3, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=4, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, "H3")

    # ---------------- Catering ----------------
    ws = wb.create_sheet("Catering")
    _sheet_title(ws, "Catering Requests")
    _header(ws, 3, ["Request No.", "Created", "Event Date", "Package", "People", "Location", "Estimate", "Payment Status", "Status", "Customer", "Email", "Phone"])
    for i,r in enumerate(catering,4):
        vals=[_value(r, "request_number"),_value(r, "created_at"),_value(r, "event_date"),_value(r, "package_type"),_value(r, "people_count"),_value(r, "location"),_value(r, "estimated_amount"),_value(r, "payment_status"),_value(r, "status"),f'{_value(r, "first_name")} {_value(r, "last_name")}',_value(r, "email"),_value(r, "phone")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _currency_columns(ws,[7])
    _finish_sheet(ws)

    # ---------------- Packages ----------------
    ws = wb.create_sheet("Packages")
    _sheet_title(ws, "Weekly Packages")
    _header(ws, 3, ["Package No.", "Created", "Days/Week", "Package Price", "Total", "Location", "Pickup Time", "Payment Method", "Payment Status", "Status", "Customer", "Email", "Phone"])
    for i,r in enumerate(packages,4):
        vals=[_value(r, "package_number"),_value(r, "created_at"),_value(r, "frequency"),_value(r, "weekly_price"),_value(r, "total_amount"),_value(r, "pickup_location"),_value(r, "pickup_time"),_value(r, "payment_method"),_value(r, "payment_status"),_value(r, "status"),f'{_value(r, "first_name")} {_value(r, "last_name")}',_value(r, "email"),_value(r, "phone")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _currency_columns(ws,[4,5])
    _finish_sheet(ws)

    # ---------------- Customers ----------------
    ws = wb.create_sheet("Customers")
    _sheet_title(ws, "Customers")
    _header(ws, 3, ["ID", "First Name", "Last Name", "Email", "Phone", "Privacy Consent", "Terms Accepted", "Marketing Consent", "Notes", "Created"])
    for i,r in enumerate(customers,4):
        vals=[_value(r, "id"),_value(r, "first_name"),_value(r, "last_name"),_value(r, "email"),_value(r, "phone"),_value(r, "privacy_consent_at"),_value(r, "terms_accepted_at"),_value(r, "marketing_consent"),_value(r, "notes"),_value(r, "created_at")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _finish_sheet(ws)

    # ---------------- Menu ----------------
    ws = wb.create_sheet("Menu")
    _sheet_title(ws, "Menu")
    _header(ws, 3, ["Date", "Day", "German Day", "Meal", "Price", "Active"])
    for i,r in enumerate(menu,4):
        vals=[_value(r, "menu_date"),_value(r, "day_name"),_value(r, "day_name_german"),_value(r, "meal_name"),_value(r, "price"),_value(r, "is_active")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _currency_columns(ws,[5])
    _finish_sheet(ws)

    # ---------------- Pickup ----------------
    ws = wb.create_sheet("Pickup Slots")
    _sheet_title(ws, "Pickup Slots")
    _header(ws, 3, ["Date", "Location", "Start", "End", "Maximum", "Current", "Active"])
    for i,r in enumerate(pickup,4):
        vals=[_value(r, "pickup_date"),_value(r, "location"),_value(r, "start_time"),_value(r, "end_time"),_value(r, "maximum_orders"),_value(r, "current_orders"),_value(r, "is_active")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _finish_sheet(ws)

    # ---------------- Allergens ----------------
    ws = wb.create_sheet("Allergens")
    _sheet_title(ws, "Menu Allergens", "Verified allergen declarations entered by the administrator")
    _header(ws, 3, ["Date", "Day", "Meal", "Menu Item", "Allergen Code", "Allergen", "Notes"])
    for i, r in enumerate(allergens, 4):
        code = str(_value(r, "allergen_code") or "")
        vals = [_value(r, "menu_date"), _value(r, "day_name"), _value(r, "meal_name"), _value(r, "item_name"), code, {
            "gluten":"Gluten-containing cereals", "milk":"Milk", "peanuts":"Peanuts", "nuts":"Nuts",
            "soy":"Soybeans", "sesame":"Sesame", "mustard":"Mustard", "celery":"Celery",
            "egg":"Eggs", "fish":"Fish", "crustaceans":"Crustaceans", "molluscs":"Molluscs",
            "lupin":"Lupin", "sulphites":"Sulphites"
        }.get(code, code), _value(r, "notes")]
        for c,v in enumerate(vals,1): ws.cell(i,c,v)
    _finish_sheet(ws)

    # ---------------- Business Settings ----------------
    ws = wb.create_sheet("Business Settings")
    _sheet_title(ws, "Business Settings", "Administrative information configured in the website")
    _header(ws, 3, ["Field", "Value"])
    if settings:
        skip = {"id", "updated_at"}
        for i, key in enumerate([k for k in settings[0].keys() if k not in skip], 4):
            ws.cell(i,1,key.replace("_"," ").title())
            ws.cell(i,2,settings[0][key])
    _finish_sheet(ws, widths={"A":28,"B":60})

    # Workbook metadata
    wb.properties.title = "Radhe Tiffin Business Workbook"
    wb.properties.subject = "Orders, revenue, expenses, salaries, inventory usage and operational P&L"
    wb.properties.creator = "Radhe Tiffin website"
    wb.save(output)
    return output
