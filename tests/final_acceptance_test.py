"""Radhe Tiffin final end-to-end acceptance test.

Run from the project directory:
    python tests/final_acceptance_test.py

The test uses a temporary SQLite database and never modifies the production DB.
It validates the individual customer order flow, customer allergy request,
dish-specific allergen matching, weekly package allergy persistence,
catering allergy persistence/admin visibility, admin order dashboard,
customer records, finance totals, Excel export, payment/status transitions,
pickup capacity release, and the health endpoint.
"""

from __future__ import annotations

import os
import re
import sys
import tempfile
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]
os.chdir(PROJECT_ROOT)

# If a project-local virtual environment exists, automatically relaunch the
# test with it. This prevents a common Windows mistake where `python` points
# to the global interpreter even though dependencies were installed in .venv.
if os.environ.get("RADHE_ACCEPTANCE_REEXEC") != "1":
    if os.name == "nt":
        local_python = PROJECT_ROOT / ".venv" / "Scripts" / "python.exe"
    else:
        local_python = PROJECT_ROOT / ".venv" / "bin" / "python"
    if local_python.exists() and Path(sys.executable).resolve() != local_python.resolve():
        env = os.environ.copy()
        env["RADHE_ACCEPTANCE_REEXEC"] = "1"
        os.execve(str(local_python), [str(local_python), str(Path(__file__).resolve())], env)

# Configure an isolated test database before importing app.py because the app
# initializes its database during module import.
test_root = Path(tempfile.mkdtemp(prefix="radhe_tiffin_acceptance_"))
os.environ["RADHE_DATABASE_PATH"] = str(test_root / "database.db")
os.environ["RADHE_ADMIN_USERNAME"] = "admin"
os.environ["RADHE_ADMIN_PASSWORD"] = "acceptance-test-password"
os.environ["PUBLIC_BASE_URL"] = "http://127.0.0.1:5000"

sys.path.insert(0, str(PROJECT_ROOT))

from app import app  # noqa: E402
from database import get_connection  # noqa: E402


app.config.update(TESTING=True)


def assert_ok(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def main() -> None:
    client = app.test_client()

    # ------------------------------------------------------------
    # 1. Customer website exposes the allergy section.
    # ------------------------------------------------------------
    response = client.get("/order/thursday")
    assert_ok(response.status_code == 200, "Customer order page did not load")
    html = response.get_data(as_text=True)
    assert_ok(
        "Do you have any food allergies or intolerances?" in html,
        "Customer allergy section is missing",
    )
    assert_ok(
        'name="customer_allergens"' in html,
        "Customer allergen controls are missing",
    )

    # ------------------------------------------------------------
    # 2. Prepare a verified dish-level allergen scenario.
    # ------------------------------------------------------------
    connection = get_connection()
    menu = connection.execute(
        """
        SELECT id, meal_name
        FROM menu
        WHERE menu_date = '2026-08-20' AND is_active = 1
        ORDER BY id DESC LIMIT 1
        """
    ).fetchone()
    assert_ok(menu is not None, "Thursday sample menu is missing")

    menu_items = connection.execute(
        """
        SELECT id, item_name
        FROM menu_items
        WHERE menu_id = ?
        ORDER BY id
        """,
        (menu["id"],),
    ).fetchall()

    target_dishes = {"2 Rotis", "Fresh Salad"}
    target_ids = {
        row["id"] for row in menu_items if row["item_name"] in target_dishes
    }
    assert_ok(
        target_ids == {
            row["id"] for row in menu_items if row["item_name"] in target_dishes
        },
        "Expected test dishes are missing",
    )
    assert_ok(len(target_ids) == 2, "Expected exactly two test dishes")

    for item_id in target_ids:
        connection.execute(
            """
            INSERT OR IGNORE INTO menu_item_allergens(menu_item_id, allergen_code)
            VALUES (?, 'milk')
            """,
            (item_id,),
        )

    slot = connection.execute(
        """
        SELECT id, current_orders, maximum_orders
        FROM pickup_slots
        WHERE pickup_date = '2026-08-20'
          AND location = 'Neu-Ulm'
          AND start_time = '14:30'
          AND is_active = 1
        LIMIT 1
        """
    ).fetchone()
    assert_ok(slot is not None, "Expected Neu-Ulm pickup slot is missing")
    before_capacity = slot["current_orders"]
    connection.commit()
    connection.close()

    # ------------------------------------------------------------
    # 3. Customer places a complete cash order with Milk selected.
    # ------------------------------------------------------------
    order_form = {
        "first_name": "Acceptance",
        "last_name": "Test",
        "email": "acceptance@example.com",
        "phone": "+491234567890",
        "quantity": "1",
        "pickup_slot_id": str(slot["id"]),
        "payment_method": "cash",
        "privacy_consent": "1",
        "terms_accepted": "1",
        "marketing_consent": "0",
        "customer_allergens": "milk",
    }
    response = client.post("/order/thursday/place-order", data=order_form)
    assert_ok(response.status_code == 200, "Customer order failed")
    confirmation = response.get_data(as_text=True)
    assert_ok("ORDER CONFIRMED" in confirmation, "Confirmation page missing")
    assert_ok("Milk" in confirmation, "Customer allergy request missing from confirmation")

    match = re.search(r"RT-\d{8}-\d{6}-\d+", confirmation)
    assert_ok(match is not None, "Order number was not generated")
    order_number = match.group(0)

    # ------------------------------------------------------------
    # 4. Verify database persistence and pickup capacity.
    # ------------------------------------------------------------
    connection = get_connection()
    order = connection.execute(
        "SELECT * FROM orders WHERE order_number = ?", (order_number,)
    ).fetchone()
    assert_ok(order is not None, "Order was not stored in the database")

    requested = connection.execute(
        """
        SELECT allergen_code
        FROM order_allergen_requests
        WHERE order_id = ?
        ORDER BY id
        """,
        (order["id"],),
    ).fetchall()
    assert_ok(
        [row["allergen_code"] for row in requested] == ["milk"],
        "Customer allergen request was not persisted correctly",
    )

    updated_slot = connection.execute(
        "SELECT current_orders FROM pickup_slots WHERE id = ?",
        (slot["id"],),
    ).fetchone()
    assert_ok(
        updated_slot["current_orders"] == before_capacity + 1,
        "Pickup capacity did not increase after order",
    )
    order_id = order["id"]
    connection.close()

    # ------------------------------------------------------------
    # 5. Admin login and order dashboard.
    # ------------------------------------------------------------
    response = client.post(
        "/admin/login",
        data={"username": "admin", "password": "acceptance-test-password"},
        follow_redirects=True,
    )
    assert_ok(response.status_code == 200, "Admin login failed")

    response = client.get("/admin?date=2026-08-20")
    assert_ok(response.status_code == 200, "Admin order dashboard failed")
    admin_html = response.get_data(as_text=True)

    for required in (
        "CUSTOMER-REPORTED ALLERGY / INTOLERANCE",
        "Milk",
        "AFFECTED DISHES — REVIEW BEFORE PREPARATION",
        "2 Rotis",
        "Fresh Salad",
    ):
        assert_ok(required in admin_html, f"Admin dashboard missing: {required}")

    # ------------------------------------------------------------
    # 6. Customer records and finance dashboard.
    # ------------------------------------------------------------
    response = client.get("/admin/customers")
    assert_ok(response.status_code == 200, "Customers page failed")
    assert_ok("Acceptance Test" in response.get_data(as_text=True), "Customer missing")

    response = client.get("/admin/finance")
    assert_ok(response.status_code == 200, "Finance page failed")
    assert_ok("€12.00" in response.get_data(as_text=True), "Finance total missing")

    # ------------------------------------------------------------
    # 7. Excel export contains the new order.
    # ------------------------------------------------------------
    response = client.get("/admin/excel/download")
    assert_ok(response.status_code == 200, "Excel download failed")
    workbook_path = test_root / "acceptance.xlsx"
    workbook_path.write_bytes(response.data)

    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    assert_ok("Orders" in workbook.sheetnames, "Orders sheet missing from Excel")
    orders_sheet = workbook["Orders"]
    rows = list(orders_sheet.values)
    assert_ok(
        any(order_number in {str(value) for value in row} for row in rows),
        "New order missing from Excel Orders sheet",
    )
    workbook.close()

    # ------------------------------------------------------------
    # 8. Weekly package: allergy request persists to every generated
    #    package-day order and package selection auto-selects weekdays.
    # ------------------------------------------------------------
    package_page = client.get("/packages")
    assert_ok(package_page.status_code == 200, "Packages page did not load")
    package_html = package_page.get_data(as_text=True)
    assert_ok('name="customer_allergens"' in package_html,
              "Package customer allergen controls are missing")
    assert_ok("entire weekly package" in package_html,
              "Package allergy scope is not clearly explained")
    assert_ok('id="packageSelector"' in package_html and "boxes.slice(0,n).forEach" in package_html,
              "Package selection does not automatically select the required weekdays")

    connection = get_connection()
    package_menu = connection.execute(
        """SELECT id, menu_date, day_name, meal_name FROM menu
           WHERE menu_date = '2026-08-18' AND is_active = 1
           ORDER BY id DESC LIMIT 1"""
    ).fetchone()
    assert_ok(package_menu is not None, "Tuesday package menu is missing")

    package_item = connection.execute(
        """SELECT id, item_name FROM menu_items
           WHERE menu_id = ? ORDER BY id LIMIT 1""",
        (package_menu["id"],),
    ).fetchone()
    assert_ok(package_item is not None, "Package menu item is missing")
    connection.execute(
        """INSERT OR IGNORE INTO menu_item_allergens(menu_item_id, allergen_code)
           VALUES (?, 'milk')""",
        (package_item["id"],),
    )

    package_slot = connection.execute(
        """SELECT id FROM pickup_slots
           WHERE pickup_date = '2026-08-18' AND location = 'Ulm'
             AND start_time = '12:00' AND end_time = '12:30'
             AND is_active = 1 LIMIT 1"""
    ).fetchone()
    assert_ok(package_slot is not None, "Tuesday package pickup slot is missing")
    connection.commit()
    connection.close()

    package_summary = client.post(
        "/packages/summary",
        data={
            "first_name": "Package", "last_name": "Acceptance",
            "email": "package@example.com", "phone": "+491234567891",
            "location": "Ulm", "pickup_time": "12:00–12:30",
            "frequency": "1", "days": "Tuesday",
            "customer_allergens": "milk",
            "privacy_consent": "1", "terms_accepted": "1",
            "marketing_consent": "0",
        },
    )
    assert_ok(package_summary.status_code == 200, "Weekly package summary failed")
    package_summary_html = package_summary.get_data(as_text=True)
    assert_ok("Milk" in package_summary_html and "entire weekly package" in package_summary_html,
              "Package summary did not display the customer allergy request")

    package_order = client.post(
        "/packages/place-order",
        data={
            "first_name": "Package", "last_name": "Acceptance",
            "email": "package@example.com", "phone": "+491234567891",
            "location": "Ulm", "pickup_start": "12:00", "pickup_end": "12:30",
            "frequency": "1", "days": "Tuesday", "customer_allergens": "milk",
            "privacy_consent": "1", "terms_accepted": "1",
            "marketing_consent": "0", "payment_method": "cash",
        },
    )
    assert_ok(package_order.status_code == 200, "Weekly package order failed")
    assert_ok("Milk" in package_order.get_data(as_text=True),
              "Package confirmation did not show the customer allergy request")

    connection = get_connection()
    weekly_package = connection.execute(
        """SELECT id, package_number FROM weekly_packages
           WHERE customer_id = (SELECT id FROM customers WHERE email = 'package@example.com'
                                ORDER BY id DESC LIMIT 1)
           ORDER BY id DESC LIMIT 1"""
    ).fetchone()
    assert_ok(weekly_package is not None, "Weekly package was not persisted")

    package_generated_order = connection.execute(
        """SELECT id, package_id FROM orders
           WHERE package_id = ? ORDER BY id DESC LIMIT 1""",
        (weekly_package["id"],),
    ).fetchone()
    assert_ok(package_generated_order is not None, "Package day order was not generated")

    package_requested = connection.execute(
        """SELECT allergen_code FROM order_allergen_requests
           WHERE order_id = ? ORDER BY id""",
        (package_generated_order["id"],),
    ).fetchall()
    assert_ok([row["allergen_code"] for row in package_requested] == ["milk"],
              "Package customer allergy was not persisted to its generated order")
    connection.close()

    # ------------------------------------------------------------
    # 9. Catering: allergy request persists and is visible to admin.
    # ------------------------------------------------------------
    catering_page = client.get("/catering")
    assert_ok(catering_page.status_code == 200, "Catering page did not load")
    catering_html = catering_page.get_data(as_text=True)
    assert_ok('name="customer_allergens"' in catering_html,
              "Catering customer allergen controls are missing")
    assert_ok("entire catering request" in catering_html,
              "Catering allergy scope is not clearly explained")

    catering_response = client.post(
        "/catering/request",
        data={
            "first_name": "Catering", "last_name": "Acceptance",
            "email": "catering@example.com", "phone": "+491234567892",
            "package_type": "office", "people_count": "20",
            "event_date": "2026-08-28", "location": "Ulm",
            "service_notes": "Acceptance test catering request",
            "customer_allergens": ["milk", "nuts"],
            "payment_method": "bank_transfer",
            "privacy_consent": "1", "terms_accepted": "1",
            "marketing_consent": "0",
        },
    )
    assert_ok(catering_response.status_code == 200, "Catering request failed")
    catering_confirmation = catering_response.get_data(as_text=True)
    assert_ok("Milk" in catering_confirmation and "Nuts" in catering_confirmation,
              "Catering confirmation did not show the customer allergy request")

    connection = get_connection()
    catering_request = connection.execute(
        """SELECT id, request_number FROM catering_requests
           WHERE customer_id = (SELECT id FROM customers WHERE email = 'catering@example.com'
                                ORDER BY id DESC LIMIT 1)
           ORDER BY id DESC LIMIT 1"""
    ).fetchone()
    assert_ok(catering_request is not None, "Catering request was not persisted")

    catering_requested = connection.execute(
        """SELECT allergen_code FROM catering_allergen_requests
           WHERE catering_request_id = ? ORDER BY allergen_code""",
        (catering_request["id"],),
    ).fetchall()
    assert_ok([row["allergen_code"] for row in catering_requested] == ["milk", "nuts"],
              "Catering customer allergies were not persisted correctly")
    connection.close()

    admin_catering = client.get("/admin/catering")
    assert_ok(admin_catering.status_code == 200, "Admin catering page did not load")
    admin_catering_html = admin_catering.get_data(as_text=True)
    assert_ok("CUSTOMER-REPORTED ALLERGY / INTOLERANCE" in admin_catering_html,
              "Admin catering allergy warning is missing")
    assert_ok("Milk" in admin_catering_html and "Nuts" in admin_catering_html,
              "Admin catering page did not display the requested allergens")

    # ------------------------------------------------------------
    # 10. Payment/status transitions.
    # ------------------------------------------------------------
    response = client.post(
        f"/admin/order/{order_id}/payment",
        data={"payment_status": "paid"},
        follow_redirects=True,
    )
    assert_ok(response.status_code == 200, "Mark-paid action failed")

    response = client.post(
        f"/admin/order/{order_id}/status",
        data={"status": "ready"},
        follow_redirects=True,
    )
    assert_ok(response.status_code == 200, "Mark-ready action failed")

    connection = get_connection()
    current = connection.execute(
        "SELECT status, payment_status FROM orders WHERE id = ?", (order_id,)
    ).fetchone()
    connection.close()
    assert_ok(current["status"] == "ready", "Ready status did not persist")
    assert_ok(current["payment_status"] == "paid", "Paid status did not persist")

    # ------------------------------------------------------------
    # 11. Cancellation releases pickup capacity.
    # ------------------------------------------------------------
    response = client.post(
        f"/admin/order/{order_id}/status",
        data={"status": "cancelled"},
        follow_redirects=True,
    )
    assert_ok(response.status_code == 200, "Cancel-order action failed")

    connection = get_connection()
    current_capacity = connection.execute(
        "SELECT current_orders FROM pickup_slots WHERE id = ?", (slot["id"],)
    ).fetchone()["current_orders"]
    connection.close()
    assert_ok(
        current_capacity == before_capacity,
        "Cancelled order did not release pickup capacity",
    )

    # ------------------------------------------------------------
    # 12. Production health endpoint.
    # ------------------------------------------------------------
    response = client.get("/health")
    assert_ok(response.status_code == 200, "Health endpoint failed")
    assert_ok(response.get_json() == {"status": "ok"}, "Health response is incorrect")

    print("ACCEPTANCE TEST PASSED")
    print(f"Order tested: {order_number}")
    print("Customer → Order → Allergy → Packages → Catering → Admin → Finance → Excel → Status → Capacity: OK")


if __name__ == "__main__":
    main()
